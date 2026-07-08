import sqlite3, json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter

DB='/tmp/operations_trending.db'; OUT='/tmp/BMG_GlazeGrade_by_Period.xlsx'
con=sqlite3.connect(DB); con.row_factory=sqlite3.Row; cur=con.cursor()
NAVY="1E2761"
HDR=Font(name='Arial',bold=True,color='FFFFFF',size=10); HDRFILL=PatternFill('solid',fgColor=NAVY)
BASE=Font(name='Arial',size=10); BOLD=Font(name='Arial',bold=True,size=10)
TITLE=Font(name='Arial',bold=True,size=14,color=NAVY); SUB=Font(name='Arial',size=9,color='666666')
CTR=Alignment('center',vertical='center'); LFT=Alignment('left',vertical='center')
thin=Side(style='thin',color='D9D9D9'); BORD=Border(thin,thin,thin,thin)
def hdrrow(ws,n,row=1):
    for c in range(1,n+1):
        x=ws.cell(row=row,column=c); x.font=HDR; x.fill=HDRFILL; x.alignment=CTR; x.border=BORD
def cscale(ws,rng):
    ws.conditional_formatting.add(rng, ColorScaleRule(start_type='num',start_value=1,start_color='F8696B',
        mid_type='num',mid_value=3,mid_color='FFEB84',end_type='num',end_value=5,end_color='63BE7B'))

master={r['pc']:r for r in cur.execute("SELECT * FROM store_master")}
gg=cur.execute("SELECT pc,period,gg,n_metrics FROM glazegrade_period").fetchall()
periods=sorted({r['period'] for r in gg})
gmap={(r['pc'],r['period']):r['gg'] for r in gg if r['gg'] is not None}
L3M=['2026-02','2026-03','2026-04']
wb=Workbook()

# ---- MAIN: GlazeGrade by Store by Period ----
ws=wb.active; ws.title='GlazeGrade by Period'
ws['A1']='GlazeGrade by Store by Period'; ws['A1'].font=TITLE
ws['A2']='Weighted 1–5 operations score (1=worst, 5=best). Recomputed from Hub v08, clipped to each store’s open→close window. Use the filter arrows on row 5 to filter by DM, Market, or Format.'; ws['A2'].font=SUB
ws['A3']='2026-05 is partial (SOS + Window Time only). L3M avg = Feb–Apr 2026.'; ws['A3'].font=SUB
top=5; fixed=['PC','Store','Market','DM','Format','L3M Avg']; H=fixed+periods
for j,h in enumerate(H,1): ws.cell(row=top,column=j,value=h)
hdrrow(ws,len(H),top)
ordered=sorted(master.values(), key=lambda m:(m['dm'],m['market'],m['name']))
for i,m in enumerate(ordered, top+1):
    pc=m['pc']
    for col,val in [(1,pc),(2,m['name']),(3,m['market']),(4,m['dm']),(5,m['format'])]:
        ws.cell(row=i,column=col,value=val)
    l3=[gmap[(pc,p)] for p in L3M if (pc,p) in gmap]
    c=ws.cell(row=i,column=6,value=round(sum(l3)/len(l3),2) if l3 else None); c.font=BOLD; c.number_format='0.00'
    for j,p in enumerate(periods,7):
        ws.cell(row=i,column=j,value=gmap.get((pc,p))).number_format='0.00'
    for j in range(1,len(H)+1):
        cc=ws.cell(row=i,column=j); cc.border=BORD; cc.alignment=LFT if j==2 else CTR
        if cc.font.bold is False: cc.font=BASE
        if j==6: cc.font=BOLD
nrows=len(ordered); last=get_column_letter(len(H))
cscale(ws, f"F{top+1}:{last}{top+nrows}")
ws.auto_filter.ref=f"A{top}:{last}{top+nrows}"
ws.freeze_panes='G6'
ws.column_dimensions['A'].width=9; ws.column_dimensions['B'].width=22
for cl in ['C','D','E']: ws.column_dimensions[cl].width=11
ws.column_dimensions['F'].width=9
for j in range(7,len(H)+1): ws.column_dimensions[get_column_letter(j)].width=8

# ---- DM SUMMARY: DM x period average store GlazeGrade ----
ws=wb.create_sheet('DM Summary')
ws['A1']='GlazeGrade by DM by Period'; ws['A1'].font=TITLE
ws['A2']='Each cell = simple average of that DM’s store GlazeGrades for the period (equal weight per store). Sorted worst→best by L3M avg.'; ws['A2'].font=SUB
top=4; H=['DM','# Stores','L3M Avg']+periods
for j,h in enumerate(H,1): ws.cell(row=top,column=j,value=h)
hdrrow(ws,len(H),top)
dms=sorted({m['dm'] for m in master.values()})
dm_pcs={dm:[pc for pc,m in master.items() if m['dm']==dm] for dm in dms}
def dm_avg(pcs,plist):
    vals=[gmap[(pc,p)] for pc in pcs for p in plist if (pc,p) in gmap]
    return round(sum(vals)/len(vals),2) if vals else None
rows=[]
for dm in dms:
    pcs=dm_pcs[dm]
    rows.append((dm,len(pcs),dm_avg(pcs,L3M),[dm_avg(pcs,[p]) for p in periods]))
rows.sort(key=lambda r:(r[2] if r[2] is not None else 99))
for i,(dm,nst,l3,pv) in enumerate(rows, top+1):
    ws.cell(row=i,column=1,value=dm).font=BOLD
    ws.cell(row=i,column=2,value=nst)
    c=ws.cell(row=i,column=3,value=l3); c.font=BOLD; c.number_format='0.00'
    for j,v in enumerate(pv,4): ws.cell(row=i,column=j,value=v).number_format='0.00'
    for j in range(1,len(H)+1):
        cc=ws.cell(row=i,column=j); cc.border=BORD; cc.alignment=LFT if j==1 else CTR
        if j not in (1,3): cc.font=BASE
last=get_column_letter(len(H))
cscale(ws, f"C{top+1}:{last}{top+len(rows)}")
ws.freeze_panes='D5'; ws.auto_filter.ref=f"A{top}:{last}{top+len(rows)}"
ws.column_dimensions['A'].width=13; ws.column_dimensions['B'].width=9; ws.column_dimensions['C'].width=9
for j in range(4,len(H)+1): ws.column_dimensions[get_column_letter(j)].width=8

# ---- RAW + RANKINGS ----
ws=wb.create_sheet('Raw Data + Rankings')
H=['PC','Store','Market','DM','Period','Metric','Raw Value','Points (1-5)']
for j,h in enumerate(H,1): ws.cell(row=1,column=j,value=h)
hdrrow(ws,len(H))
rws=cur.execute("""SELECT o.pc,m.name,m.market,m.dm,o.period,o.metric,o.value,o.points
   FROM ops_metric_raw o JOIN store_master m ON m.pc=o.pc ORDER BY m.dm,m.name,o.period,o.metric""").fetchall()
for i,r in enumerate(rws,2):
    for j,v in enumerate(r,1):
        x=ws.cell(row=i,column=j,value=(round(v,4) if isinstance(v,float) else v)); x.font=BASE; x.border=BORD
        x.alignment=LFT if j<=6 else CTR
for cl,w in zip('ABCDEFGH',[9,22,11,11,10,16,12,12]): ws.column_dimensions[cl].width=w
ws.freeze_panes='A2'; ws.auto_filter.ref=f"A1:H{len(rws)+1}"

# ---- WEIGHTING & THRESHOLDS ----
ws=wb.create_sheet('Weighting & Thresholds')
ws['A1']='GlazeGrade Metric Definitions — Thresholds & Weights'; ws['A1'].font=TITLE
ws['A2']='Weights renormalize over metrics present each period. v1 = currently applied (matches Hub). v2 = proposed rebalance (not applied).'; ws['A2'].font=SUB
top=4; H=['Metric','Direction','5 pts','4 pts','3 pts','2 pts','1 pt','Wt DT (v1)','Wt Inline (v1)','Wt DT (v2)','Wt Inline (v2)','Status','Note']
for j,h in enumerate(H,1): ws.cell(row=top,column=j,value=h)
hdrrow(ws,len(H),top)
def bins(direction,cuts):
    if not cuts: return ['—']*5
    a,b,c,d=cuts
    if direction=='high_good': return [f'≥{d}',f'{c}–{d}',f'{b}–{c}',f'{a}–{b}',f'<{a}']
    return [f'<{a}',f'{a}–{b}',f'{b}–{c}',f'{c}–{d}',f'≥{d}']
for i,r in enumerate(cur.execute("SELECT * FROM metric_definition").fetchall(), top+1):
    cuts=json.loads(r['cuts']) if r['cuts'] else None
    b5,b4,b3,b2,b1=bins(r['direction'],cuts)
    status='PLACEHOLDER' if r['placeholder'] else ('active' if r['active'] else 'inactive')
    vals=[r['name'],r['direction'] or '—',b5,b4,b3,b2,b1,r['weight_dt_v1'],r['weight_in_v1'],r['weight_dt_v2'],r['weight_in_v2'],status,r['note'] or '']
    for j,v in enumerate(vals,1):
        x=ws.cell(row=i,column=j,value=v); x.font=BASE; x.border=BORD
        if 8<=j<=11 and isinstance(v,(int,float)): x.number_format='0.0%'
        x.alignment=CTR if 2<=j<=11 else LFT
        if status=='PLACEHOLDER': x.fill=PatternFill('solid',fgColor='FFF2CC')
for cl,w in zip('ABCDEFGHIJKLM',[15,11,8,8,8,8,8,11,12,11,12,12,40]): ws.column_dimensions[cl].width=w
ws.freeze_panes='A5'

# ---- README ----
ws=wb.create_sheet('README'); ws.sheet_view.showGridLines=False
ws['A1']='BMG GlazeGrade — by Store by Period'; ws['A1'].font=TITLE
for i,(a,b) in enumerate([
 ('Source','Recomputed from BMG Ops Metrics Hub v08; clipped to each store’s open→close window (Bluemont store listing). DM names from Store Contact List.'),
 ('GlazeGrade by Period','MAIN tab. Store × period, color-scaled 1–5. Filter arrows on row 5 → filter by DM / Market / Format. L3M Avg = Feb–Apr 2026.'),
 ('DM Summary','GlazeGrade rolled up by DM by period (simple avg of the DM’s stores). Sorted worst→best by L3M avg.'),
 ('Raw Data + Rankings','Every store × period × metric: raw value and its 1–5 points. Filterable.'),
 ('Weighting & Thresholds','Scoring rules: bin cutoffs + DT/Inline weights (v1 applied, v2 proposed).'),
 ('Method','Metric binned 1–5 by threshold; GlazeGrade = weighted avg by store format, renormalized over metrics present.'),
 ('Caveat','2026-05 partial (SOS + Window Time only).'),
],start=3):
    ws.cell(row=i,column=1,value=a).font=BOLD; ws.cell(row=i,column=2,value=b).font=BASE
ws.column_dimensions['A'].width=22; ws.column_dimensions['B'].width=100
wb.move_sheet('README', -(len(wb.sheetnames)-1))
wb.save(OUT)
print('saved | stores',nrows,'| periods',len(periods),'| DMs',len(rows))
con.close()
