"""
generate_outputs.py — Regenerate the dashboard HTML + Excel export FROM operations_trending.db.
This is the update loop: refresh the DB, run this, get fresh deliverables.

Usage: python generate_outputs.py <db> <template.html> <out_html> <out_xlsx>
"""
import json, sqlite3, sys, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DB, TPL, OUT_HTML, OUT_XLSX = sys.argv[1:5]
METRICS=['SOS','Window Time','OSAT','Accuracy','Complaints/10K','Cert SL/AGM','Training','Hourly TO','COGS Var','Labor Var']
con=sqlite3.connect(DB); con.row_factory=sqlite3.Row; cur=con.cursor()

# ---- reconstruct v12-shaped DATA dict ----
stores=[]
for sm in cur.execute("SELECT * FROM store_master ORDER BY pc").fetchall():
    pc=sm['pc']
    sn=cur.execute("SELECT * FROM store_snapshot WHERE pc=?",(pc,)).fetchone()
    metrics={}; mvals={}
    for r in cur.execute("SELECT metric,score,raw_value FROM metric_score WHERE pc=?",(pc,)):
        metrics[r['metric']]=r['score']; mvals[r['metric']]=r['raw_value']
    trend=[{'period':r['period'],'gg':r['gg'],'sss':r['sss'],'drag':r['drag'],'pf_sss':r['pf_sss']}
           for r in cur.execute("SELECT * FROM store_period WHERE pc=? ORDER BY period",(pc,))]
    events=[{'new_store':r['new_store'],'event_type':r['event_type'],'date_opened':r['date_opened'],
             'impact':r['impact'],'net':r['net'],'src':r['src']}
            for r in cur.execute("SELECT * FROM cc_event WHERE pc=?",(pc,))]
    gmtl=[{'gm':r['gm'],'start':r['start'],'end':r['end']}
          for r in cur.execute("SELECT * FROM gm_timeline WHERE pc=? ORDER BY start",(pc,))]
    # per-period raw values for each metric (source: ops_metric_monthly) -> {metric:[{period,raw}]}
    metric_trend={}
    for r in cur.execute("SELECT metric,period,raw_value FROM ops_metric_monthly WHERE pc=? ORDER BY period",(pc,)):
        if r['raw_value'] is None: continue
        metric_trend.setdefault(r['metric'],[]).append({'period':r['period'],'raw':r['raw_value']})
    stores.append({'pc':pc,'name':sm['name'],'market':sm['market'],'dm':sm['dm'],'rd':sm['rd'],
        'format':sm['format'],'actual_sss':sn['actual_sss'],'baseline_sss':sn['baseline_sss'],
        'net_drag':sn['net_drag'],'new_gg':sn['new_gg'],'old_gg':sn['old_gg'],'sales_4w':sn['sales_4w'],
        'annual_sales':sn['annual_sales'],'n_events':sn['n_events'],'weakest':sn['weakest'],
        'current_gm':sn['current_gm'],'gm_changes':sn['gm_changes'],'gm_tenure':sn['gm_tenure'],
        'metrics':metrics,'trend':trend,'events':events,'ext_count':sn['ext_count'],
        'metric_values':mvals,'gm_timeline':gmtl,'category':sn['category'],'lat':sm['lat'],'lon':sm['lon'],
        'metric_trend':metric_trend})
DATA={'stores':stores}
# metric definitions (direction / cuts / unit) for front-end scoring + coloring
try:
    _mdpath=os.path.join(os.path.dirname(TPL) or '.','metric_definitions.json')
    if not os.path.exists(_mdpath):
        _mdpath=os.path.join(os.path.dirname(os.path.abspath(__file__)),'metric_definitions.json')
    _md=json.load(open(_mdpath))
    DATA['metric_defs']={m['name']:{'direction':m.get('direction'),'cuts':m.get('cuts'),'unit':m.get('unit')}
                         for m in _md.get('metrics',[]) if m.get('active') and m.get('cuts')}
except Exception as _e:
    DATA['metric_defs']={}
try:
    _cols=[c[0] for c in cur.execute('SELECT * FROM ninebox LIMIT 0').description]
    DATA['ninebox']=[dict(zip(_cols,row)) for row in cur.execute('SELECT * FROM ninebox').fetchall()]
except Exception as _e:
    DATA['ninebox']=[]


# ---- HTML: inject DATA into template ----
tpl=open(TPL).read()
open(OUT_HTML,'w').write(tpl.replace('__DATA_JSON__', json.dumps(DATA)))

# ---- EXCEL ----
NAVY="1E2761"
HDR=Font(name='Arial',bold=True,color='FFFFFF',size=10); HDRFILL=PatternFill('solid',fgColor=NAVY)
BASE=Font(name='Arial',size=10); BOLD=Font(name='Arial',bold=True,size=10)
TITLE=Font(name='Arial',bold=True,size=14,color=NAVY); SUB=Font(name='Arial',size=9,color='666666')
CTR=Alignment('center',vertical='center'); LFT=Alignment('left',vertical='center')
thin=Side(style='thin',color='D9D9D9'); BORD=Border(thin,thin,thin,thin)
PCT='0.0%;(0.0%);-'; PCT2='0.00%;(0.00%);-'; CUR='$#,##0;($#,##0);-'; NUM2='0.00'
wb=Workbook()
def hdr(ws,n,row=1):
    for c in range(1,n+1):
        x=ws.cell(row=row,column=c); x.font=HDR; x.fill=HDRFILL; x.alignment=CTR; x.border=BORD
def tbl(ws,H,rows,F,W,start=1):
    for j,h in enumerate(H,1): ws.cell(row=start,column=j,value=h)
    hdr(ws,len(H),start)
    for i,r in enumerate(rows,start+1):
        for j,v in enumerate(r,1):
            x=ws.cell(row=i,column=j,value=v); x.font=BASE; x.border=BORD; f=F[j-1]
            if f: x.number_format=f
            x.alignment=LFT if (f is None or f=='@') else CTR
    for j,w in enumerate(W,1): ws.column_dimensions[get_column_letter(j)].width=w
    ws.freeze_panes=ws.cell(row=start+1,column=1)

ws=wb.active; ws.title='README'; ws.sheet_view.showGridLines=False
ws['A1']='BMG Operations Trending — Data Export'; ws['A1'].font=TITLE
ws['A2']='Generated from operations_trending.db'; ws['A2'].font=SUB
for i,(a,b) in enumerate([('Store Summary','SSS actual vs proforma baseline, net drag, guest growth, sales, GM, category.'),
    ('Ops Metrics','10 ops metrics: 1-5 score (1=worst,5=best) + raw value.'),
    ('Monthly Trend','Per store-month: Guest Growth, SSS, C/C drag, proforma SSS.'),
    ('External Events','C/C / new-store events with net SSS impact.')],start=4):
    ws.cell(row=i,column=1,value=a).font=BOLD; ws.cell(row=i,column=2,value=b).font=BASE
ws.column_dimensions['A'].width=22; ws.column_dimensions['B'].width=80

ws=wb.create_sheet('Store Summary')
H=['Store #','Store Name','Market','DM','RD','Format','Category','Actual SSS','Baseline SSS','Net Drag','Ext Events','New GG','Old GG','Sales (4wk)','Annual Sales','Current GM','GM Changes','GM Tenure (yr)','Weakest Metrics']
F=['@','@','@','@','@','@','@',PCT,PCT,PCT,'0','0.00','0.00',CUR,CUR,'@','0','0.0','@']
W=[9,22,11,11,11,8,18,11,12,10,9,9,9,13,14,20,11,13,30]
rows=[[s['pc'],s['name'],s['market'],s['dm'],s['rd'],s['format'],s['category'],s['actual_sss'],s['baseline_sss'],s['net_drag'],s['ext_count'],s['new_gg'],s['old_gg'],s['sales_4w'],s['annual_sales'],s['current_gm'],s['gm_changes'],s['gm_tenure'],s['weakest']] for s in stores]
tbl(ws,H,rows,F,W); tr=len(rows)+2
ws.cell(row=tr,column=2,value='PORTFOLIO').font=BOLD
for col in (14,15):
    c=ws.cell(row=tr,column=col,value=f'=SUM({get_column_letter(col)}2:{get_column_letter(col)}{len(rows)+1})'); c.font=BOLD; c.number_format=CUR

ws=wb.create_sheet('Ops Metrics')
H=['Store #','Store Name','Market','DM']+[f'{m} (score)' for m in METRICS]+[f'{m} (raw)' for m in METRICS]
F=['@','@','@','@']+['0.0']*10+[None]*10; W=[9,22,11,11]+[13]*20
rows=[[s['pc'],s['name'],s['market'],s['dm']]+[s['metrics'].get(m) for m in METRICS]+[round(s['metric_values'].get(m),4) if isinstance(s['metric_values'].get(m),(int,float)) else None for m in METRICS] for s in stores]
tbl(ws,H,rows,F,W)

ws=wb.create_sheet('Monthly Trend')
H=['Store #','Store Name','Market','DM','Period','Guest Growth','SSS','C/C Drag','Proforma SSS']
F=['@','@','@','@','@',NUM2,PCT2,PCT2,PCT2]; W=[9,22,11,11,10,13,11,11,13]
rows=[[s['pc'],s['name'],s['market'],s['dm'],t['period'],t['gg'],t['sss'],t['drag'],t['pf_sss']] for s in stores for t in s['trend']]
tbl(ws,H,rows,F,W)

ws=wb.create_sheet('External Events')
H=['Affected Store #','Affected Store','Market','DM','New/Event Store','Event Type','Date Opened','Impact','Net SSS Impact','Source']
F=['@','@','@','@','@','@','@',PCT2,PCT2,'@']; W=[13,22,11,11,22,12,12,11,14,20]
rows=[[s['pc'],s['name'],s['market'],s['dm'],e['new_store'],e['event_type'],e['date_opened'],e['impact'],e['net'],e['src']] for s in stores for e in s['events']]
tbl(ws,H,rows,F,W)
wb.save(OUT_XLSX)
print('HTML  ->',OUT_HTML)
print('XLSX  ->',OUT_XLSX)
print('stores regenerated:',len(stores))
con.close()
